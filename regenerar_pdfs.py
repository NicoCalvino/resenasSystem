"""
regenerar_pdfs.py — Regenera los PDFs a partir de un archivo Excel generado previamente.

No requiere acceso a las APIs de Rappi/PedidosYa ni a los CSVs de Mercado Libre.
Basta con el archivo Excel (.xlsx) que produce el sistema normalmente.

Uso (línea de comandos):
    python regenerar_pdfs.py resenas_2026-04-13.xlsx
    python regenerar_pdfs.py resenas_2026-04-13.xlsx --output ./mis_pdfs
    python regenerar_pdfs.py resenas_2026-04-13.xlsx --grupo Billinghurst

Uso (como módulo desde gui.py):
    from regenerar_pdfs import regenerar_desde_excel
    regenerar_desde_excel("ruta/al/excel.xlsx", output_dir="./informes",
                          log_fn=print, grupo_filtro=None)
"""

import os
import re
import sys
import argparse
from pathlib import Path
from datetime import datetime
from collections import OrderedDict, defaultdict

# Asegurar que el módulo report/ sea importable
sys.path.insert(0, str(Path(__file__).parent))


def _cargar_env_archivo() -> None:
    """
    Carga las variables del archivo .env del proyecto en os.environ.

    Es necesario porque cuando este módulo se invoca desde la GUI (que NO
    llama a load_dotenv()), las variables como PEDIDOS_SHEETS_URL no están
    disponibles y los PDFs mensuales salen con 0 pedidos.

    Usa python-dotenv si está instalado; si no, hace un parseo simple.
    No sobreescribe variables ya definidas en el entorno.
    """
    env_path = Path(__file__).parent / ".env"
    if not env_path.exists():
        return
    try:
        from dotenv import load_dotenv
        load_dotenv(dotenv_path=env_path, override=False)
        return
    except ImportError:
        pass
    # Fallback manual
    try:
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            k = k.strip()
            v = v.strip().strip('"').strip("'")
            if k and k not in os.environ:
                os.environ[k] = v
    except Exception:
        pass


# ── Lectura de hojas ──────────────────────────────────────────────────────────

def _leer_resenas_desde_hoja(ws) -> list[dict]:
    """
    Lee todas las reseñas de la hoja unificada "Resenas".
    Las filas de datos empiezan en la fila 3 (fila 1 = título, fila 2 = encabezados).
    Columnas: Fecha y Hora | Nro Orden | Aplicacion | Grupo/Local | Marca |
              Estrellas | Plato | Etiquetas | Comentario
    """
    resenas = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        orden_id = row[1]
        if not orden_id:
            continue  # fila vacía o fin de datos

        # Fecha (puede ser string o datetime según openpyxl)
        fecha_raw = row[0]
        fecha = None
        if fecha_raw:
            if isinstance(fecha_raw, datetime):
                fecha = fecha_raw
            else:
                for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S"):
                    try:
                        fecha = datetime.strptime(str(fecha_raw), fmt)
                        break
                    except ValueError:
                        pass

        # Etiquetas: separadas por " | "
        etiquetas_raw = str(row[7]).strip() if row[7] else ""
        tags = [t.strip() for t in etiquetas_raw.split("|") if t.strip()] if etiquetas_raw else []

        resenas.append({
            "fecha":        fecha,
            "orden_id":     str(orden_id),
            "app":          str(row[2]) if row[2] else "",
            "grupo_local":  str(row[3]) if row[3] else "",
            "marca":        str(row[4]) if row[4] else "",
            "estrellas":    int(row[5]) if row[5] else 1,
            "plato":        str(row[6]) if row[6] else "",
            "tags":         tags,
            "comentario":   str(row[8]) if row[8] else "",
            "grave":        False,   # ya no se almacena en Resenas; se recalcula si es necesario
        })

    return resenas


def _leer_resenas_desde_hoja_legacy(ws, app: str) -> list[dict]:
    """
    Compatibilidad con Excels generados antes del cambio de esquema.
    Esquema antiguo (3 hojas separadas por app):
    Fecha y Hora | Nro Orden | Grupo/Local | Marca | Estrellas | Plato | Etiquetas | Comentario | Error Grave
    """
    resenas = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        orden_id = row[1]
        if not orden_id:
            continue
        fecha_raw = row[0]
        fecha = None
        if fecha_raw:
            if isinstance(fecha_raw, datetime):
                fecha = fecha_raw
            else:
                for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S"):
                    try:
                        fecha = datetime.strptime(str(fecha_raw), fmt)
                        break
                    except ValueError:
                        pass
        etiquetas_raw = str(row[6]).strip() if row[6] else ""
        tags = [t.strip() for t in etiquetas_raw.split("|") if t.strip()] if etiquetas_raw else []
        grave_raw = str(row[8]).strip().upper() if len(row) > 8 and row[8] else ""
        grave = grave_raw in ("SÍ", "SI", "SÌ", "S")
        resenas.append({
            "fecha":       fecha,
            "orden_id":    str(orden_id),
            "app":         app,
            "grupo_local": str(row[2]) if row[2] else "",
            "marca":       str(row[3]) if row[3] else "",
            "estrellas":   int(row[4]) if row[4] else 1,
            "plato":       str(row[5]) if row[5] else "",
            "tags":        tags,
            "comentario":  str(row[7]) if row[7] else "",
            "grave":       grave,
        })
    return resenas


def _leer_reclamos_desde_hoja(ws) -> list[dict]:
    """
    Lee todos los reclamos de la hoja "Reclamos".
    Columnas: Fecha y Hora | Nro Orden | App | Grupo/Local | Marca |
              Platos Pedidos | Platos Reclamados | Motivo | Comentario del cliente
    """
    reclamos = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        orden_id = row[1]
        if not orden_id:
            continue

        fecha_raw = row[0]
        fecha = None
        if fecha_raw:
            if isinstance(fecha_raw, datetime):
                fecha = fecha_raw
            else:
                for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S"):
                    try:
                        fecha = datetime.strptime(str(fecha_raw), fmt)
                        break
                    except ValueError:
                        pass

        reclamos.append({
            "fecha":             fecha,
            "orden_id":          str(orden_id),
            "app":               str(row[2]) if row[2] else "",
            "grupo_local":       str(row[3]) if row[3] else "",
            "marca":             str(row[4]) if row[4] else "",
            "platos_pedidos":    str(row[5]) if row[5] else "",
            "platos_reclamados": str(row[6]) if row[6] else "",
            "razon":             str(row[7]) if row[7] else "",
            "comentario":        str(row[8]) if row[8] else "",
        })

    return reclamos


def _leer_totales_desde_hoja(ws) -> dict[str, dict]:
    """
    Lee los totales por grupo desde la hoja "Totales".
    Columnas fijas: Grupo/Local | Rappi | PedidosYa | Mercado Libre | Total Ordenes
    Columnas variables: una por cada marca (Green Eat, Las Gracias, Tea Connection, etc.)
    Columnas finales: Reclamos | Resenas Negativas | Errores Graves | % Error

    Usa la fila de encabezados para ubicar las columnas de forma robusta,
    independientemente de cuántas marcas estén presentes.
    """
    # Leer encabezados de la fila 2 para ubicar las columnas dinámicamente
    headers = [cell.value for cell in ws[2]]
    def _col(name):
        """Devuelve el índice (0-based) de la columna con ese nombre, o None."""
        for i, h in enumerate(headers):
            if h and str(h).strip().lower() == name.strip().lower():
                return i
        return None

    idx_total    = _col("Total Ordenes")   or 4
    idx_neg      = _col("Resenas Negativas")
    idx_graves   = _col("Errores Graves")

    # Fallback por si los encabezados tienen tildes (excel generado antes de este cambio)
    if idx_neg is None:
        idx_neg    = _col("Reseñas Negativas") or _col("Resenas Negativas")
    if idx_graves is None:
        idx_graves = _col("Errores Graves")

    def _int(v):
        try:
            return int(float(v)) if v is not None else 0
        except (ValueError, TypeError):
            return 0

    totales = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        grupo = row[0]
        if not grupo or str(grupo).strip().upper() == "TOTAL":
            continue

        totales[str(grupo).strip()] = {
            "total_ordenes":     _int(row[idx_total])                      if idx_total  is not None else 0,
            "resenas_negativas": _int(row[idx_neg])   if idx_neg    is not None else 0,
            "errores_graves":    _int(row[idx_graves]) if idx_graves is not None else 0,
        }

    return totales


# ── Adaptadores para build_report() ──────────────────────────────────────────

def _adaptar_resenas_para_pdf(resenas: list[dict]) -> list[dict]:
    """
    Convierte la lista plana de reseñas al formato de órdenes agrupadas
    que espera build_report() (múltiples platos por orden_id).
    """
    por_orden: OrderedDict = OrderedDict()
    for r in resenas:
        oid = r["orden_id"]
        if oid not in por_orden:
            por_orden[oid] = {
                "fecha":    r["fecha"].strftime("%d/%m/%Y %H:%M") if r["fecha"] else "",
                "orden":    oid,
                "marca":    r["marca"],
                "app":      r["app"],
                "estrellas": r["estrellas"],
                "platos":   [],
            }
        por_orden[oid]["platos"].append({
            "nombre":     r["plato"],
            "tags":       r["tags"],
            "comentario": r["comentario"],
            "grave":      r["grave"],
        })
    return list(por_orden.values())


def _adaptar_reclamos_para_pdf(reclamos: list[dict]) -> list[dict]:
    """
    Convierte la lista de reclamos al formato que espera build_report().
    """
    ordenados = sorted(reclamos, key=lambda r: r["fecha"] or datetime.min)
    return [
        {
            "fecha":             rc["fecha"].strftime("%d/%m/%Y %H:%M") if rc["fecha"] else "",
            "orden":             rc["orden_id"],
            "app":               rc["app"],
            "marca":             rc["marca"],
            "platos_pedidos":    rc["platos_pedidos"],
            "platos_reclamados": rc["platos_reclamados"],
            "razon":             rc["razon"],
            "comentario":        rc["comentario"],
        }
        for rc in ordenados
    ]


# ── Función principal ─────────────────────────────────────────────────────────

def regenerar_desde_excel(
    ruta_excel: str,
    output_dir: str = "./informes",
    log_fn=None,
    grupo_filtro: str = None,
) -> tuple[int, int]:
    """
    Lee el Excel y genera un PDF por cada grupo/local encontrado.

    Args:
        ruta_excel:    Ruta al archivo .xlsx generado por el sistema.
        output_dir:    Carpeta donde se guardarán los PDFs.
        log_fn:        Función de logging; recibe (mensaje, tag) donde
                       tag ∈ {"info","ok","warn","error","head"}.
                       Si es None se usa print().
        grupo_filtro:  Si se especifica, solo genera el PDF de ese grupo.

    Returns:
        (pdfs_generados, total_grupos)
    """
    from openpyxl import load_workbook
    from report.generador_pdf import build_report

    def log(msg, tag="info"):
        if log_fn:
            log_fn(msg, tag)
        else:
            prefijos = {"ok": "✓", "error": "✗", "warn": "!", "head": "─"}
            print(f"{prefijos.get(tag, ' ')} {msg}")

    ruta_excel = Path(ruta_excel)
    if not ruta_excel.exists():
        log(f"No se encontró el archivo: {ruta_excel}", "error")
        return 0, 0

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    log(f"Leyendo Excel: {ruta_excel.name}", "head")

    try:
        wb = load_workbook(ruta_excel, data_only=True)
    except Exception as e:
        log(f"No se pudo abrir el Excel: {e}", "error")
        return 0, 0

    # ── Leer reseñas de la hoja unificada "Resenas" ──────────────────────────
    todas_resenas: list[dict] = []
    hoja_resenas = next((n for n in wb.sheetnames if n.lower() in ("resenas", "reseñas")), None)
    if hoja_resenas:
        todas_resenas = _leer_resenas_desde_hoja(wb[hoja_resenas])
        log(f"Resenas: {len(todas_resenas)} filas", "info")
    else:
        # Compatibilidad con Excels generados con el esquema anterior (3 hojas por app)
        for app in ("Rappi", "PedidosYa", "Mercado Libre"):
            if app in wb.sheetnames:
                rs = _leer_resenas_desde_hoja_legacy(wb[app], app)
                todas_resenas.extend(rs)
                log(f"{app} (legacy): {len(rs)} resenas", "info")

    # ── Leer reclamos ─────────────────────────────────────────────────────────
    todos_reclamos: list[dict] = []
    if "Reclamos" in wb.sheetnames:
        todos_reclamos = _leer_reclamos_desde_hoja(wb["Reclamos"])
        log(f"Reclamos: {len(todos_reclamos)}", "info")

    # ── Leer totales ──────────────────────────────────────────────────────────
    totales: dict[str, dict] = {}
    if "Totales" in wb.sheetnames:
        totales = _leer_totales_desde_hoja(wb["Totales"])
        log(f"Totales: {len(totales)} grupos", "info")
    else:
        log("Hoja 'Totales' no encontrada — los indicadores estarán en 0", "warn")

    # ── Inferir fecha del nombre del archivo ─────────────────────────────────
    # Acepta cualquier nombre que contenga YYYY-MM-DD, por ejemplo:
    #   resenas_2026-04-13.xlsx
    #   informe_2026-05-15_143330.xlsx
    stem = ruta_excel.stem
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", stem)
    fecha_pdf = None
    if m:
        try:
            fecha_pdf = datetime.strptime(m.group(0), "%Y-%m-%d")
        except ValueError:
            fecha_pdf = None
    if fecha_pdf is None:
        fecha_pdf = datetime.now()
        log(f"No se pudo inferir la fecha del nombre '{stem}'; usando hoy.", "warn")

    fecha_display = f"{fecha_pdf.day}/{fecha_pdf.month}/{fecha_pdf.year}"

    # ── Agrupar por Grupo/Local ───────────────────────────────────────────────
    resenas_por_grupo: dict[str, list] = defaultdict(list)
    for r in todas_resenas:
        resenas_por_grupo[r["grupo_local"]].append(r)

    reclamos_por_grupo: dict[str, list] = defaultdict(list)
    for rc in todos_reclamos:
        reclamos_por_grupo[rc["grupo_local"]].append(rc)

    # Universo de grupos: union de los que tienen resenas + los de totales
    todos_grupos = sorted(set(resenas_por_grupo.keys()) | set(totales.keys()))

    if grupo_filtro:
        todos_grupos = [g for g in todos_grupos if g == grupo_filtro]
        if not todos_grupos:
            log(f"Grupo '{grupo_filtro}' no encontrado en el Excel.", "error")
            return 0, 0

    log(f"Generando {len(todos_grupos)} PDFs...", "head")

    pdfs_generados = 0
    for grupo in todos_grupos:
        resenas_g  = resenas_por_grupo.get(grupo, [])
        reclamos_g = reclamos_por_grupo.get(grupo, [])
        totales_g  = totales.get(grupo, {
            "total_ordenes": 0,
            "resenas_negativas": 0,
            "errores_graves": 0,
        })

        data = {
            "local":             grupo,
            "fecha":             fecha_display,
            "total_ordenes":     totales_g["total_ordenes"],
            "resenas_negativas": totales_g["resenas_negativas"],
            "errores_graves":    totales_g["errores_graves"],
            "resenas":           _adaptar_resenas_para_pdf(resenas_g),
            "reclamos":          _adaptar_reclamos_para_pdf(reclamos_g),
        }

        nombre_pdf = f"{grupo}_{fecha_pdf.strftime('%Y-%m-%d')}.pdf"
        nombre_pdf = "".join(
            c if c.isalnum() or c in "._- " else "_" for c in nombre_pdf)
        ruta_pdf = output_path / nombre_pdf

        try:
            build_report(data, str(ruta_pdf))
            pdfs_generados += 1
            log(f"PDF generado: {nombre_pdf}", "ok")
        except Exception as e:
            log(f"Error generando PDF para '{grupo}': {e}", "error")

    log(f"Completado: {pdfs_generados}/{len(todos_grupos)} PDFs generados en {output_path.resolve()}", "head")
    return pdfs_generados, len(todos_grupos)


# ── PDFs mensuales (historico.json) ──────────────────────────────────────────

def regenerar_pdfs_mensuales(
    output_dir: str = "./informes",
    anio: int = None,
    mes: int = None,
    log_fn=None,
    fecha_hasta=None,
) -> tuple[int, int]:
    """
    Genera los 6 PDFs mensuales de reclamos por grupo (marca × DK)
    a partir del historico.json compartido en Google Drive.

    Args:
        output_dir:   Carpeta donde se guardarán los PDFs.
        anio:         Año del mes a reportar. Si es None se infiere de
                      fecha_hasta o, en su defecto, del mes actual.
        mes:          Mes del mes a reportar. Mismo criterio que anio.
        log_fn:       Función de logging; recibe (mensaje, tag).
        fecha_hasta:  Último día a incluir en el acumulado.
                      Puede ser datetime, date o string "YYYY-MM-DD".
                      Default: hoy.

    Returns:
        (pdfs_generados, total_grupos)
    """
    # Cargar .env: imprescindible cuando se llama desde la GUI, que no
    # invoca load_dotenv() — sin esto, PEDIDOS_SHEETS_URL queda vacía y
    # los PDFs mensuales se generan con 0 pedidos.
    _cargar_env_archivo()

    from report.generador_pdf_mensual import generar_pdfs_mensuales

    def log(msg, tag="info"):
        if log_fn:
            log_fn(msg, tag)
        else:
            prefijos = {"ok": "✓", "error": "✗", "warn": "!", "head": "─"}
            print(f"{prefijos.get(tag, ' ')} {msg}")

    # ── Normalizar fecha_hasta ───────────────────────────────────────────────
    # Acepta datetime, date o str "YYYY-MM-DD" (formato que usa la GUI).
    if isinstance(fecha_hasta, str):
        try:
            fecha_hasta = datetime.strptime(fecha_hasta.strip(), "%Y-%m-%d")
        except ValueError:
            log(f"Fecha hasta inválida: {fecha_hasta!r} — se usará hoy.", "warn")
            fecha_hasta = None
    elif fecha_hasta is not None and not isinstance(fecha_hasta, datetime):
        # date u objeto similar
        try:
            fecha_hasta = datetime(fecha_hasta.year, fecha_hasta.month, fecha_hasta.day)
        except Exception:
            log(f"Fecha hasta no reconocida: {fecha_hasta!r} — se usará hoy.", "warn")
            fecha_hasta = None

    hoy = datetime.now()
    if fecha_hasta is None:
        fecha_hasta = hoy

    # Si no se pasaron anio/mes, derivarlos de fecha_hasta para que la regeneración
    # use el mes de "hasta" (la GUI suele pedir un período del mes en curso).
    anio = anio or fecha_hasta.year
    mes  = mes  or fecha_hasta.month

    if not os.environ.get("PEDIDOS_SHEETS_URL", "").strip():
        log("PEDIDOS_SHEETS_URL no está definida — los PDFs saldrán con 0 pedidos.", "warn")

    log(f"Generando PDFs mensuales — {mes:02d}/{anio} (hasta {fecha_hasta.strftime('%Y-%m-%d')})", "head")

    try:
        pdfs = generar_pdfs_mensuales(
            output_dir=output_dir,
            anio=anio,
            mes=mes,
            fecha_hasta=fecha_hasta,
        )
        generados = len(pdfs)
        for p in pdfs:
            log(f"PDF generado: {Path(p).name}", "ok")
        log(f"Completado: {generados} PDFs mensuales generados en {Path(output_dir).resolve()}", "head")
        return generados, 6
    except Exception as e:
        log(f"Error generando PDFs mensuales: {e}", "error")
        return 0, 6


# ── Entrada por línea de comandos ─────────────────────────────────────────────

def _parse_args():
    p = argparse.ArgumentParser(
        description="Regenera PDFs de resenas a partir de un Excel existente.")
    p.add_argument("excel", nargs="?", default=None,
                   help="Ruta al archivo Excel (.xlsx) generado por el sistema")
    p.add_argument("--output", default="./informes",
                   help="Carpeta de salida de los PDFs (default: ./informes)")
    p.add_argument("--grupo", default=None,
                   help="Nombre exacto del grupo a regenerar (opcional; default: todos)")
    p.add_argument("--mensuales", action="store_true",
                   help="Genera los 6 PDFs mensuales desde historico.json en lugar del Excel")
    p.add_argument("--anio", type=int, default=None,
                   help="Año para los PDFs mensuales (default: año actual)")
    p.add_argument("--mes", type=int, default=None,
                   help="Mes para los PDFs mensuales (default: mes actual)")
    return p.parse_args()


if __name__ == "__main__":
    args = _parse_args()
    if args.mensuales:
        ok, total = regenerar_pdfs_mensuales(
            output_dir=args.output,
            anio=args.anio,
            mes=args.mes,
        )
    else:
        if not args.excel:
            print("Error: se requiere el archivo Excel (o usar --mensuales para los PDFs mensuales).")
            sys.exit(1)
        ok, total = regenerar_desde_excel(
            ruta_excel=args.excel,
            output_dir=args.output,
            grupo_filtro=args.grupo,
        )
    sys.exit(0 if ok == total else 1)
