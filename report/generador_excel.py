"""
Generador de Excel con reporte completo del proceso.
Pestanas: Reclamos | Resenas | Totales
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from config.models import Resena, Reclamo, ResumenLocal

# Colores
COLOR_GRAVE_BG    = "FFE0E0"
COLOR_HEADER_TOT  = "1A1A2E"
COLOR_SUBHEAD_TOT = "16213E"
COLOR_ALT_ROW     = "F8F8F8"

_thin   = Side(style="thin", color="CCCCCC")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _header_style(bg, fg="FFFFFF", bold=True, size=10):
    font  = Font(name="Arial", bold=bold, color=fg, size=size)
    fill  = PatternFill("solid", start_color=bg)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return font, fill, align


def _set_row(ws, row_idx, values, font=None, fill=None, align=None, border=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        if font:   cell.font      = font
        if fill:   cell.fill      = fill
        if align:  cell.alignment = align
        if border: cell.border    = border


# ── Pestana de Resenas (todas las apps combinadas) ───────────────────────────

COLS_RESENAS = [
    "Fecha y Hora", "Nro Orden", "Aplicacion", "Grupo/Local", "Marca",
    "Estrellas", "Plato", "Etiquetas", "Comentario"
]

def _escribir_hoja_resenas(ws, resenas):
    color_bg = "2C3E50"
    color_fg = "FFFFFF"
    font_h, fill_h, align_h = _header_style(color_bg, color_fg, size=10)

    ws.merge_cells("A1:" + get_column_letter(len(COLS_RESENAS)) + "1")
    t = ws["A1"]
    t.value     = "Resenas negativas (1-2 estrellas) - todas las aplicaciones"
    t.font      = Font(name="Arial", bold=True, size=12, color=color_fg)
    t.fill      = PatternFill("solid", start_color=color_bg)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    _set_row(ws, 2, COLS_RESENAS, font=font_h, fill=fill_h, align=align_h, border=_border)
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "A3"

    font_d   = Font(name="Arial", size=9)
    align_d  = Alignment(vertical="center", wrap_text=True)
    fill_alt = PatternFill("solid", start_color=COLOR_ALT_ROW)

    resenas_ord = sorted(resenas, key=lambda r: (r.app, r.local_nombre, r.fecha_orden or datetime.min))

    for i, r in enumerate(resenas_ord, 3):
        fill_row = fill_alt if i % 2 == 0 else None
        values = [
            r.fecha_orden.strftime("%d/%m/%Y %H:%M") if r.fecha_orden else "",
            r.orden_id,
            r.app,
            r.local_nombre,
            r.marca,
            r.estrellas,
            r.plato,
            " | ".join(r.tags) if r.tags else "",
            r.comentario,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font      = font_d
            cell.alignment = align_d
            cell.border    = _border
            if fill_row:
                cell.fill = fill_row
        ws.row_dimensions[i].height = 15

    anchos = [16, 16, 14, 16, 16, 9, 28, 28, 45]
    for col, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    return len(resenas_ord)


# ── Pestana de Reclamos ──────────────────────────────────────────────────────

COLS_RECLAMOS = [
    "Fecha y Hora", "Nro Orden", "App", "Grupo/Local", "Marca",
    "Platos Pedidos", "Platos Reclamados", "Motivo", "Comentario del cliente", "Error Grave"
]

def _escribir_hoja_reclamos(ws, reclamos):
    from processor.procesador import es_error_grave as _es_grave

    color_bg = "444444"
    color_fg = "FFFFFF"
    font_h, fill_h, align_h = _header_style(color_bg, color_fg, size=10)

    ws.merge_cells("A1:" + get_column_letter(len(COLS_RECLAMOS)) + "1")
    t = ws["A1"]
    t.value     = "Reclamos / Compensaciones"
    t.font      = Font(name="Arial", bold=True, size=12, color=color_fg)
    t.fill      = PatternFill("solid", start_color=color_bg)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    _set_row(ws, 2, COLS_RECLAMOS, font=font_h, fill=fill_h, align=align_h, border=_border)
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "A3"

    font_d    = Font(name="Arial", size=9)
    font_g    = Font(name="Arial", size=9, color="CC0000", bold=True)
    align_d   = Alignment(vertical="center", wrap_text=True)
    fill_alt  = PatternFill("solid", start_color=COLOR_ALT_ROW)
    fill_grave = PatternFill("solid", start_color=COLOR_GRAVE_BG)

    reclamos_ord = sorted(reclamos, key=lambda r: (r.app, r.local_nombre, r.fecha_orden))

    for i, r in enumerate(reclamos_ord, 3):
        grave    = _es_grave(r.comentario or "", [r.razon or ""])
        fill_row = fill_grave if grave else (fill_alt if i % 2 == 0 else None)
        font_row = font_g if grave else font_d

        values = [
            r.fecha_orden.strftime("%d/%m/%Y %H:%M") if r.fecha_orden else "",
            r.orden_id,
            r.app,
            r.local_nombre,
            r.marca,
            r.platos_pedidos,
            r.platos_reclamados,
            r.razon,
            r.comentario,
            "SI" if grave else "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font      = font_row
            cell.alignment = align_d
            cell.border    = _border
            if fill_row:
                cell.fill = fill_row
        ws.row_dimensions[i].height = 15

    anchos = [16, 16, 12, 16, 16, 40, 30, 36, 45, 12]
    for col, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    return len(reclamos_ord)


# ── Pestana de Totales ───────────────────────────────────────────────────────

def _escribir_hoja_totales(ws, resumenes, totales_rappi, totales_peya,
                            totales_ml, reclamos=None,
                            totales_marca_por_grupo=None):
    totales_marca_por_grupo = totales_marca_por_grupo or {}

    todas_marcas = sorted(set(
        marca
        for marcas in totales_marca_por_grupo.values()
        for marca in marcas
    ))

    # Columnas:
    # 1 Grupo | 2 Rappi | 3 PedidosYa | 4 ML | 5 Total
    # 6..5+n  marca_1..marca_n
    # 5+n+1 Reclamos | 5+n+2 Res.Neg | 5+n+3 Err.Graves | 5+n+4 % Error
    n_marcas     = len(todas_marcas)
    col_reclamos = 6 + n_marcas
    col_graves   = 8 + n_marcas
    col_pct      = 9 + n_marcas
    n_cols       = col_pct

    ws.merge_cells("A1:" + get_column_letter(n_cols) + "1")
    c = ws["A1"]
    c.value     = "Totales de ordenes por grupo y aplicacion"
    c.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill      = PatternFill("solid", start_color=COLOR_HEADER_TOT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    cols = (
        ["Grupo/Local", "Rappi", "PedidosYa", "Mercado Libre", "Total Ordenes"]
        + todas_marcas
        + ["Reclamos", "Resenas Negativas", "Errores Graves", "% Error"]
    )
    font_h, fill_h, align_h = _header_style(COLOR_SUBHEAD_TOT, "FFFFFF", size=10)
    _set_row(ws, 2, cols, font=font_h, fill=fill_h, align=align_h, border=_border)
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "A3"

    grupos = sorted(set(
        list(totales_rappi.keys()) +
        list(totales_peya.keys()) +
        list(totales_ml.keys()) +
        [r.local_nombre for r in resumenes]
    ))

    mapa = {r.local_nombre: r for r in resumenes}

    reclamos_por_grupo = {}
    for rec in (reclamos or []):
        reclamos_por_grupo[rec.local_nombre] = reclamos_por_grupo.get(rec.local_nombre, 0) + 1

    font_d   = Font(name="Arial", size=9)
    font_b   = Font(name="Arial", size=9, bold=True)
    align_c  = Alignment(horizontal="center", vertical="center")
    align_l  = Alignment(horizontal="left",   vertical="center")
    fill_alt = PatternFill("solid", start_color=COLOR_ALT_ROW)

    for i, grupo in enumerate(grupos, 3):
        fill_row = fill_alt if i % 2 == 0 else None

        r_rappi = totales_rappi.get(grupo, 0)
        r_peya  = totales_peya.get(grupo, 0)
        r_ml    = totales_ml.get(grupo, 0)
        total   = r_rappi + r_peya + r_ml

        resumen    = mapa.get(grupo)
        neg        = resumen.resenas_negativas if resumen else 0
        graves     = resumen.errores_graves    if resumen else 0
        n_reclamos = reclamos_por_grupo.get(grupo, 0)

        marcas_grupo = totales_marca_por_grupo.get(grupo, {})
        brand_vals   = [marcas_grupo.get(m, 0) for m in todas_marcas]

        col_rec_l    = get_column_letter(col_reclamos)
        col_graves_l = get_column_letter(col_graves)
        pct_formula  = "=IFERROR(" + col_graves_l + str(i) + "/" + col_rec_l + str(i) + ",0)"

        row_vals = (
            [grupo, r_rappi, r_peya, r_ml, total]
            + brand_vals
            + [n_reclamos, neg, graves, pct_formula]
        )

        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = _border
            if fill_row:
                cell.fill = fill_row
            if col == 1:
                cell.font      = font_b
                cell.alignment = align_l
            else:
                cell.font      = font_d
                cell.alignment = align_c

        ws.cell(row=i, column=col_pct).number_format = "0.0%"
        ws.row_dimensions[i].height = 15

    last    = len(grupos) + 2
    tot_row = last + 1
    ws.cell(row=tot_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
    for col in range(2, col_pct + 1):
        letter = get_column_letter(col)
        cell = ws.cell(row=tot_row, column=col,
                       value="=SUM(" + letter + "3:" + letter + str(last) + ")")
        cell.font      = Font(name="Arial", bold=True, size=10)
        cell.fill      = PatternFill("solid", start_color="E8E8E8")
        cell.border    = _border
        cell.alignment = Alignment(horizontal="center")
    ws.cell(row=tot_row, column=col_pct).number_format = "0.0%"

    base_widths  = [22, 12, 12, 14, 14]
    brand_widths = [16] * n_marcas
    tail_widths  = [12, 18, 14, 10]
    all_widths   = base_widths + brand_widths + tail_widths
    for col, w in enumerate(all_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ── Funcion principal ────────────────────────────────────────────────────────

def generar_excel(
    resumenes,
    todas_resenas,
    totales_rappi,
    totales_peya,
    totales_ml,
    ruta_salida,
    reclamos=None,
    totales_marca_por_grupo=None,
):
    wb = Workbook()
    wb.remove(wb.active)

    if reclamos:
        ws_rec = wb.create_sheet(title="Reclamos")
        _escribir_hoja_reclamos(ws_rec, reclamos)

    ws_res = wb.create_sheet(title="Resenas")
    _escribir_hoja_resenas(ws_res, todas_resenas)

    ws_tot = wb.create_sheet(title="Totales")
    _escribir_hoja_totales(
        ws_tot, resumenes,
        totales_rappi, totales_peya, totales_ml,
        reclamos, totales_marca_por_grupo,
    )

    wb.save(ruta_salida)
    return ruta_salida
